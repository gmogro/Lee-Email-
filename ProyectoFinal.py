import tkinter as tk
import poplib
from email.parser import BytesParser, Parser
from openpyxl import Workbook

class VentanaLogin:
    def __init__(self, master):
     self.master = master
     self.frame = tk.Frame(self.master)
     self.master.title("----------Login-------------")
     self.master.geometry("320x150+500+250")
     self.master.resizable(width="FALSE", height="FALSE")

     self.lbUser = tk.Label(self.frame, text="User")
     self.lbUser.pack()
     self.input_User = tk.Entry(self.frame, width = 35)
     self.input_User.pack()
     self.lbPass = tk.Label(self.frame, text="Password")
     self.lbPass.pack()
     self.input_Pass = tk.Entry(self.frame, width = 35, show="*")
     self.input_Pass.pack()

     self.button1 = tk.Button(self.frame, text = 'Login', width = 25, command=self._login_btn_clicked)
     self.button1.pack()
     self.frame.pack()

    def _login_btn_clicked(self):
        # print("Clicked")
        username = self.input_User.get()
        password = self.input_Pass.get()

        print(username, password)

        #conexion a servicios de gmail
        M = poplib.POP3_SSL('pop.gmail.com')
        M.user(username)
        M.pass_(password)
        #obtiene el numero de mensaje
        numero = len(M.list()[1])
        #Obtiene mensaje
        global response, headerLines, bytes
        for i in range(numero):
            # Se lee el mensaje
            response, headerLines, bytes = M.retr(i + 1)
        #se mete todo en un string
        mensaje = b'\n'.join(headerLines)
        #se parsea
        # Se parsea el mensaje
        p = BytesParser()
        email = p.parsebytes(mensaje)
        #crea nueva ventana
        self.new_window(email)


    def new_window(self,lista):
     self.newWindow = tk.Toplevel(self.master)
     self.app = VentanaLista(self.newWindow,lista)

class VentanaLista:

    def __init__(self, master,lista):

        self.master = master
        self.frame = tk.Frame(master)
        master.title("Lista")
        master.geometry("320x150+500+250")

        self.frame.grid(column = 0 ,row =0 , padx=(50,50),pady=(10,10))
        self.frame.columnconfigure(0,weight=1)
        self.frame.rowconfigure(0,weight=1)

        self.btnOrdenar = tk.Button(self.frame, text = 'Ordenar', width = 25)
        self.btnOrdenar.grid(row=1, column=1)
        self.btnReporte = tk.Button(self.frame, text = 'Reporte', width = 25,command=self.Reporte(lista))
        self.btnReporte.grid(row=1, column=2)
        self.btnExcel = tk.Button(self.frame, text='Excel', width=25, command=self.Excel(lista))
        self.btnExcel.grid(row=1, column=3)

        self.lbFrom = tk.Label(self.frame, text="From")
        self.lbFrom.grid(row=2, column=1)
        self.lbTo = tk.Label(self.frame, text="To")
        self.lbTo.grid(row=2, column=2)
        self.lbSubject = tk.Label(self.frame, text="Subject")
        self.lbSubject.grid(row=2, column=3)
        self.lbID = tk.Label(self.frame, text="ID")
        self.lbID.grid(row=2, column=4)

        numero = len(lista)
        for row in range(3,numero+2,1):
            self.From = tk.Entry(self.frame, width = 50)
            self.From.insert(0,lista["From"])
            self.From.grid(row=row, column=1)
            self.To = tk.Entry(self.frame, width=25)
            self.To.insert(0, lista["To"])
            self.To.grid(row=row, column=2)
            self.Subject = tk.Entry(self.frame, width=35)
            self.Subject.insert(0, lista["Subject"])
            self.Subject.grid(row=row, column=3)
            self.ID = tk.Entry(self.frame, width=50)
            self.ID.insert(0, lista['message-id'])
            self.ID.grid(row=row, column=4)
            """print("From: " + lista["From"])
            print("To: " + lista["To"])
            print("Subject: " + lista["Subject"])
            print("ID: " + lista['message-id'])"""
        self.frame.pack()

    def close_windows(self):
         self.master.destroy()

    def Reporte(self,lista):
        f = open("C:/Email.txt", 'w')
        numero = len(lista)
        for l in range(numero):
            f.write("From: " + lista["From"] + " " + "To: " + lista["To"] + " " + "Subject: " + lista["Subject"] + " " + "ID: " + lista['message-id'] + '\n')

    def Excel(self,lista):
        wb = Workbook()
        ws = wb.active
        ws.title = "Email"
        numero = len(lista)
        for l in range(2,numero):
            d = ws.cell(row=l,column=2)
            d.value = lista["From"]
            d = ws.cell(row=l, column=3)
            d.value = lista["To"]
            d = ws.cell(row=l, column=4)
            d.value = lista["Subject"]
            d = ws.cell(row=l, column=5)
            d.value = lista['message-id']
        wb.save('C:/email.xlsx')

def main():
    root = tk.Tk()
    app = VentanaLogin(root)
    root.mainloop()

if __name__ == '__main__':
    main()
